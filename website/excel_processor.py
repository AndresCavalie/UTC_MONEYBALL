import openpyxl
import pandas as pd
def process(excelraw):
    sheet = excelraw
    
    sheet = openpyxl.load_workbook(sheet)
    sheet = sheet.active
    df = pd.DataFrame()
    from pathlib import Path

    

    
    for row in sheet.iter_rows(2, sheet.max_row):
    
        df.loc[len(df), ['Date', 'Opp', 'Poss #', 'Tag Start', 'Poss', 'Start', 'Shot Trigger', 'Secondary', 'Usr', 'Scr', 'Psr', 'Shot', 'Result', 'Assist+', 'FTM', 'FTA', 'Open 3', 'Shooter', 'Passer', 'C&S Jumper', 'ESQ', 'ShotClock', 'P1', 'P2', 'P3', 'P4', 'P5', '1 Tag', '2 Tag', '3 Tag', '4 Tag', '5 Tag', 'OffReb', 'First Trigger', 'Bolt', 'Paint Touch Time', 'PostTouches', 'NumPasses', 'AJ', 'KH', 'JW', 'HH', 'SA', 'JJ', 'DW', 'AS', 'JT', 'KC', 'DD', 'JS', 'RB', 'GL', 'BR', 'R2', 'PM2', 'PT2', 'NP2', 'PT3', 'NP3', 'D3']] = row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, row[7].value, row[8].value, row[9].value, row[10].value, row[11].value, row[12].value, row[13].value, row[14].value, row[15].value, row[16].value, row[17].value, row[18].value, row[19].value, row[20].value, row[21].value, row[22].value, row[23].value, row[24].value, row[25].value, row[26].value, row[27].value, row[28].value, row[29].value, row[30].value, row[31].value, row[32].value, row[33].value, row[34].value, row[35].value, row[36].value, row[37].value, row[38].value, row[39].value, row[40].value, row[41].value, row[42].value, row[43].value, row[44].value, row[45].value, row[46].value, row[47].value, row[48].value, row[49].value, row[50].value, row[51].value, row[52].value, row[53].value, row[54].value, row[55].value, row[56].value, row[57].value, row[58].value, row[59].value
        
    

    return df